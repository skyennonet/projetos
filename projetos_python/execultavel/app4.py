import openpyxl

def search_term_in_workbook(term_to_search):
    term_found = False
    # Carregando arquivo
    book = openpyxl.load_workbook('imp_ms.xlsx')

    # Verificando todas as planilhas existentes no arquivo
    all_sheets = book.sheetnames

    # Iterando por todas as planilhas
    for sheet_name in all_sheets:
        sheet = book[sheet_name]

        # Iterando por todas as linhas e verificando se o termo está presente
        for row in sheet.iter_rows(max_col=8):  # Buscar até a 8ª coluna
            row_values = [cell.value for cell in row]
            if any(term_to_search.lower() in str(value).lower() for value in row_values):
                # Se o termo for encontrado em alguma célula, mostra a linha
                print(f"Termo '{term_to_search}' encontrado na planilha '{sheet_name}' - Linha:")
                print(*[value if value else "" for value in row_values])
                term_found = True

                # Opção para substituir o item encontrado
                replace_option = input("Deseja fazer alguma alteração? (S/N): ").lower()
                if replace_option == 's':
                    old_value = term_to_search
                    new_value = input("Qual termo deseja alterar? ")
                    new_value_replace = input("Insira a nova informação: ")
                    if new_value in row_values:
                        index = row_values.index(new_value)
                        row[index].value = new_value_replace

                break

    if not term_found:
        print(f"Item '{term_to_search}' não encontrado nas planilhas.")

    # Salvando as alterações no arquivo
    book.save('imp_ms.xlsx')

while True:
    option = input("Escolha uma opção:\n1 - Fazer uma busca\n2 - Finalizar consulta\n")
    if option == '1':
        term_to_search = input("Digite a informação da impressora: ")
        search_term_in_workbook(term_to_search.lower())
    elif option == '2':
        print("Ok, tenha um bom dia.")
        break
    else:
        print("Opção inválida. Escolha novamente.")
1