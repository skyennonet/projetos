import openpyxl

def search_term_in_workbook(term_to_search):
    term_found = False # essa linha inicializar a variavel term_found
    # Carregando arquivo
    book = openpyxl.load_workbook('imp_ms.xlsx')

    # Verificando todas as planilhas existentes no arquivo
    all_sheets = book.sheetnames

    # Pedindo ao usuário para digitar o termo a ser pesquisado
    term_to_search = input("Digite a informação da impressora: ")

    # Iterando por todas as planilhas
    for sheet_name in all_sheets:
        sheet = book[sheet_name]

        # Iterando por todas as linhas e verificando se o termo está presente
        for row in sheet.iter_rows(max_col=8):  # Buscar até a 8ª coluna
            row_values = [cell.value for cell in row]
            if any(term_to_search.lower() in str(value).lower() for value in row_values):
                # Se o termo for encontrado em alguma célula, mostra a linha
                print(f"Termo '{term_to_search}' encontrado na planilha '{sheet_name}' - Linha:")
                print(*[value if value else "" for value in row_values])
                term_found = True
                break

    if not term_found:
        print(f"Item '{term_to_search}' não encontrado nas planilhas.")

while True:
    option = input("escolha uma oopção: \n1 - Fazer uma busca\n2 - finalizar consulta\n")

    if option == '1':
        term_to_search = input("digite a informação da impressora: ")
        search_term_in_workbook(term_to_search.lower())
    elif option =='2':
        print("ok, tenha um bom dia.")
        break
    else:
        print("opção invalida. escolha novamente.")

