import openpyxl

#Criar uma planilha 
book = openpyxl.Workbook()
#como visualizar paginas existentes
print(book.sheetnames)
#como criar uma pagina
book.create_sheet('frutas')
#como selecionar uma pagina
frutas_page = book['frutas']
frutas_page.append(['fruta', 'quantidade', 'preço'])
frutas_page.append(['banana', '5', 'R$3,90'])
frutas_page.append(['maça', '5', 'R$3,90'])
frutas_page.append(['melancia', '5', 'R$3,90'])
frutas_page.append(['uva', '5', 'R$3,90'])
#salvar a planilha
book.save('planilha de compras.xlsx')