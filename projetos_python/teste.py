import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox
import json

# Nome do arquivo de configuração para salvar a localização do arquivo XLSX
config_file = "config.json"

# Variável para armazenar o caminho do arquivo XLSX selecionado
file_path = None

def load_config():
    global file_path
    try:
        with open(config_file, "r") as f:
            config = json.load(f)
            file_path = config.get("file_path")
    except (FileNotFoundError, json.JSONDecodeError):
        file_path = None

def save_config():
    global file_path
    config = {"file_path": file_path}
    with open(config_file, "w") as f:
        json.dump(config, f)

def search_term_in_workbook(term_to_search, new_text):
    global file_path
    term_found = False
    if file_path is None:
        messagebox.showinfo("Erro", "Selecione um arquivo XLSX antes de fazer a busca.")
        return

    # Carregando arquivo
    book = openpyxl.load_workbook(file_path)

    # Verificando todas as planilhas existentes no arquivo
    all_sheets = book.sheetnames

    # Iterando por todas as planilhas
    for sheet_name in all_sheets:
        sheet = book[sheet_name]

        # Iterando por todas as linhas e verificando se o termo está presente
        for row in sheet.iter_rows(max_col=8):  # Buscar até a 8ª coluna
            row_values = [cell.value for cell in row]
            if any(term_to_search.lower() in str(value).lower() for value in row_values):
                # Substituindo o termo pelo novo texto na linha
                for cell in row:
                    if term_to_search.lower() in str(cell.value).lower():
                        cell.value = new_text

                term_found = True

    if term_found:
        search_result_text.insert(tk.END, f"Item '{term_to_search}' substituído por '{new_text}' nas planilhas.\n")
    else:
        search_result_text.insert(tk.END, f"Item '{term_to_search}' não encontrado nas planilhas.\n")

    # Salvando as alterações no arquivo
    book.save(file_path)

def perform_search():
    global file_path
    term_to_search = search_entry.get().lower()
    search_result_text.delete(1.0, tk.END)  # Limpa o resultado da busca anterior
    search_term_in_workbook(term_to_search, None)

def replace_item():
    global file_path
    term_to_replace = search_entry.get().lower()
    new_text = new_text_entry.get()  # Obtém o novo texto do Entry correspondente
    search_result_text.delete(1.0, tk.END)  # Limpa o resultado da busca anterior
    if new_text:  # Verifica se o novo texto foi fornecido
        search_term_in_workbook(term_to_replace, new_text)
        save_config()  # Salva a configuração após a substituição
    else:
        messagebox.showinfo("Erro", "Digite o novo texto para realizar a substituição.")

def select_file():
    global file_path
    file_path = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx")])
    if file_path:
        messagebox.showinfo("Sucesso", "Arquivo XLSX selecionado com sucesso!")
        save_config()

# Carrega a configuração ao iniciar o programa
load_config()

# Criação da janela principal
root = tk.Tk()
root.title("Pesquisa e Substituição de Itens em Excel")
root.geometry("500x400")

# Componentes da interface
search_frame = tk.Frame(root)
search_frame.pack()

search_label = tk.Label(search_frame, text="Digite a informação da impressora:")
search_label.grid(row=0, column=0)

search_entry = tk.Entry(search_frame, width=30)
search_entry.grid(row=0, column=1)

search_button = tk.Button(search_frame, text="Buscar", command=perform_search)
search_button.grid(row=0, column=2)

search_result_text = tk.Text(root, wrap=tk.WORD, width=60, height=10)
search_result_text.pack()

replace_frame = tk.Frame(root)
replace_frame.pack()

new_value_label = tk.Label(replace_frame, text="Insira o texto a ser substituído:")
new_value_label.grid(row=0, column=0)

new_value_entry = tk.Entry(replace_frame, width=30)
new_value_entry.grid(row=0, column=1)

new_text_label = tk.Label(replace_frame, text="Insira o novo texto:")
new_text_label.grid(row=1, column=0)

new_text_entry = tk.Entry(replace_frame, width=30)
new_text_entry.grid(row=1, column=1)

replace_button = tk.Button(root, text="Substituir Item", command=replace_item)
replace_button.pack()

file_button = tk.Button(root, text="Selecionar Arquivo", command=select_file)
file_button.pack()

# Carrega a configuração ao iniciar o programa
load_config()

# Inicia a execução da interface
root.mainloop()
