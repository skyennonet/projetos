import openpyxl

#carregando arquivo

book= openpyxl.load_workbook('imp_ms.xlsx')
#selecionando uma planilha
#printer = book['MP 402']                

#verificando todas as planilhas existentes no arquivo
all_sheets = book.sheetnames
#interando por todas as planilhas e imprimndo os valores da 1ª a 8ª coluna
for sheet_name in all_sheets:
    sheet = book[sheet_name]
    data = []

    for rows in sheet.iter_rows(min_row=1, max_row=5, max_col=8):
        row_data = [cell.value for cell in rows]
        data.append(row_data)

    print(f"Valores da planilha {sheet_name}:")
    for row in data:
        print(*row)
    
    
