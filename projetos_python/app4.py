import openpyxl

def search_term_in_workbook(term_to_search):
    # Carregando arquivo
    book = openpyxl.load_workbook('imp_ms.xlsx')

    # Verificando todas as planilhas existentes no arquivo
    all_sheets = book.sheetnames

    # Variável para verificar se o termo foi encontrado
    term_found = False

    # Iterando por todas as planilhas
    for sheet_name in all_sheets:
        sheet = book[sheet_name]

        # Iterando por todas as linhas e verificando se o termo está presente
        for row in sheet.iter_rows(max_col=8):  # Buscar até a 8ª coluna
            row_values = [cell.value for cell in row]
            if any(term_to_search.lower() in str(value).lower() for value in row_values):
                # Se o termo for encontrado em alguma célula, mostra a linha
                print(f"Termo '{term_to_search}' encontrado na planilha '{sheet_name}' - Linha:")
                print(*[value if value else "" for value in row_values])
                term_found = True
                break

    if not term_found:
        print(f"Termo '{term_to_search}' não encontrado nas planilhas.")


while True:
    option = input("Escolha uma opção:\n1 - Buscar novamente\n2 - Finalizar consulta\n")
    
    if option == '1':
        term_to_search = input("Digite o termo a ser pesquisado: ")
        search_term_in_workbook(term_to_search)
    elif option == '2':
        print("Ok, tenha um bom dia.")
        break
    else:
        print("Opção inválida. Escolha novamente.")
